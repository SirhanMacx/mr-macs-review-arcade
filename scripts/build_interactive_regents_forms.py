#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import urlopen

import fitz
import openpyxl
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
CATALOG_PATH = ROOT / "data" / "regents-past-exam-catalog.json"
OUT = ROOT / "data" / "regents-released-practice-exams.json"
ASSET_ROOT = ROOT / "assets" / "regents-released-forms"
TMP = ROOT / ".tmp" / "interactive-regents-forms"

ESSAY_COLUMNS = [0, 0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5]
COMMON = set(
    "the a an and or but if then this that these those with from into onto about above below because while "
    "where when which what who whom whose are was were been being have has had will would could should may "
    "might can cannot one two three four five six document source passage excerpt cartoon map graph chart table "
    "author shows shown based according question answer choice directions write explain identify describe".split()
)

COURSE_META = {
    "Grade 10 Global History II": {
        "profileId": "global-history-ii",
        "assetPrefix": "global",
        "label": "Global History II",
        "shortTitle": "Part II: Document-Based CRQs",
        "essayTitle": "Part III: Document-Based Enduring Issues Essay",
        "essayPrompt": "Identify and explain an enduring issue raised by this set of documents. Argue why the issue is significant and how it has endured across time. Use evidence from at least three of the five documents and outside information.",
        "essayDocMinimum": 3,
    },
    "Grade 11 U.S. History": {
        "profileId": "us-history",
        "assetPrefix": "us",
        "label": "U.S. History",
        "shortTitle": "Part II: Document-Based Short Essays",
        "scaffoldTitle": "Part IIIA: Civic Literacy Short-Answer Questions",
        "essayTitle": "Part IIIB: Civic Literacy Essay",
        "essayPrompt": "Identify the constitutional or civic issue raised by the documents. Explain the historical circumstances surrounding the issue, describe efforts by people or government to address it, and discuss the extent to which those efforts were successful. Use evidence from the documents and outside information.",
        "essayDocMinimum": 4,
    },
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def clean(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def clean_display(value: str) -> str:
    text = clean(value)
    replacements = {
        "fi ": "fi",
        "fl ": "fl",
        "ൿ": "fi",
        "ൺ": "I",
        "—": "-",
        "–": "-",
        "“": '"',
        "”": '"',
        "‘": "'",
        "’": "'",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+([,.;:?!])", r"\1", text)
    text = re.sub(r"\(\s*([1-4])\s*\)", r"(\1)", text)
    return text


def numeric(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"\d+(?:\.5)?", text):
            return float(text) if "." in text else int(text)
    return None


def download(url: str, dest: Path) -> Path:
    if dest.exists() and dest.stat().st_size:
        return dest
    dest.parent.mkdir(parents=True, exist_ok=True)
    with urlopen(url, timeout=60) as response:
        dest.write_bytes(response.read())
    return dest


def url_filename(url: str, fallback: str) -> str:
    name = Path(urlparse(url).path).name
    return name or fallback


def extract_scoring_key(path: Path) -> dict[int, str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    keys: dict[int, str] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            for index in range(len(values) - 1):
                q = numeric(values[index])
                key = numeric(values[index + 1])
                if q is None or key is None:
                    continue
                if float(q).is_integer() and 1 <= int(q) <= 28 and int(key) in {1, 2, 3, 4}:
                    keys.setdefault(int(q), str(int(key)))
    missing = [n for n in range(1, 29) if n not in keys]
    if missing:
        raise RuntimeError(f"{path.name}: missing MCQ scoring keys {missing}")
    return keys


def extract_conversion_chart(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        header_row = None
        header_cols: list[int] = []
        headers: list[float] = []
        for row_index in range(1, ws.max_row + 1):
            nums = []
            for col_index in range(1, ws.max_column + 1):
                value = numeric(ws.cell(row_index, col_index).value)
                if value in ESSAY_COLUMNS:
                    nums.append((col_index, float(value)))
            if len(nums) >= 10 and nums[0][1] == 0 and nums[-1][1] == 5:
                header_row = row_index
                header_cols = [col for col, _ in nums[:11]]
                headers = [value for _, value in nums[:11]]
                break
        if not header_row:
            continue

        rows: dict[str, list[int]] = {}
        empty_streak = 0
        first_header_col = min(header_cols)
        for row_index in range(header_row + 1, ws.max_row + 1):
            base = None
            for col_index in range(first_header_col - 1, 0, -1):
                value = numeric(ws.cell(row_index, col_index).value)
                if value is not None and float(value).is_integer():
                    base = int(value)
                    break
            if base is None:
                if rows:
                    empty_streak += 1
                    if empty_streak >= 3:
                        break
                continue
            values = []
            for col_index in header_cols:
                value = numeric(ws.cell(row_index, col_index).value)
                values.append(None if value is None else int(value))
            if len(values) == len(headers) and all(value is not None for value in values):
                rows[str(base)] = values
                empty_streak = 0

        if rows:
            return {
                "sourceFile": path.name,
                "essayColumns": headers,
                "rows": rows,
            }
    raise RuntimeError(f"{path.name}: could not parse conversion chart")


def page_asset(pdf: fitz.Document, out_dir: Path, page_number: int, label: str) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"page-{page_number:02d}.jpg"
    if not out_path.exists():
        page = pdf[page_number - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(1.65, 1.65), alpha=False)
        image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        image.save(out_path, "JPEG", quality=72, optimize=True, progressive=True)
    return {
        "src": "../../" + out_path.relative_to(ROOT).as_posix(),
        "label": label,
        "officialPage": page_number,
    }


def page_texts(pdf: fitz.Document) -> list[str]:
    return [page.get_text("text") for page in pdf]


def find_page(texts: list[str], patterns: list[str], start: int = 1, end: int | None = None) -> int | None:
    end = end or len(texts)
    for page_number in range(max(1, start), min(end, len(texts)) + 1):
        text = texts[page_number - 1]
        if all(re.search(pattern, text, re.I) for pattern in patterns):
            return page_number
    return None


def has_doc_marker(text: str) -> bool:
    return bool(re.search(r"\bDocument\s+\d", text, re.I))


def doc_pages(texts: list[str], start: int | None, end: int | None, count: int) -> list[int]:
    start = start or 1
    end = end or len(texts) + 1
    candidates = [page for page in range(start, min(end, len(texts) + 1)) if has_doc_marker(texts[page - 1])]
    picked = candidates[:count]
    page = start
    while len(picked) < count and page < min(end, len(texts) + 1):
        if page not in picked:
            picked.append(page)
        page += 1
    return picked[:count]


def keywords(text: str, limit: int = 10) -> list[str]:
    words = re.findall(r"[A-Za-z][A-Za-z'-]{3,}", text.lower())
    counts: dict[str, int] = {}
    for word in words:
        if word in COMMON or len(word) < 5:
            continue
        counts[word] = counts.get(word, 0) + 1
    return [word for word, _ in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:limit]]


def make_doc(course: str, administration: str, section: str, doc_number: int, page_number: int, asset: dict, text: str) -> dict:
    return {
        "docNumber": doc_number,
        "source": f"{administration} {section} Document {doc_number} (official page {page_number})",
        "keywords": keywords(text),
        "stimulusImages": [asset],
    }


def question_marker(number: int) -> re.Pattern:
    # Regents PDFs often include maps with standalone electoral-vote numbers at
    # the start of extracted lines. Only treat horizontal same-line spacing as a
    # question marker so map labels like "11\nS.C." do not become Question 11.
    return re.compile(rf"(?m)^[^\S\r\n]*(?:[\u2000-\u200a\u00a0][^\S\r\n]*)?{number}(?:\t|[^\S\r\n]+(?=\S))")


def any_question_marker() -> re.Pattern:
    return re.compile(r"(?m)^[^\S\r\n]*(?:[\u2000-\u200a\u00a0][^\S\r\n]*)?(?:[1-9]|1\d|2[0-8])(?:\t|[^\S\r\n]+(?=\S))")


def page_question_markers(text: str) -> list[tuple[int, int]]:
    markers = []
    for number in range(1, 29):
        match = question_marker(number).search(text)
        if match:
            markers.append((match.start(), number))
    return sorted(markers)


def trim_choice_tail(text: str) -> str:
    stops = [
        r"\n\s*(?:[1-9]|1\d|2[0-8])[\t \u00a0\u2000-\u200a]+(?=[A-Z\"“])",
        r"\n\s*Base your answers?\b",
        r"\n\s*Part I\b",
        r"\n\s*Directions\b",
        r"\n\s*Source:\s*$",
        r"\n\s*Global Hist\.",
        r"\n\s*U\.S\. Hist\.",
        r"\n\s*\[\d+\]",
    ]
    end = len(text)
    for pattern in stops:
        match = re.search(pattern, text, flags=re.I)
        if match:
            end = min(end, match.start())
    return text[:end]


def choice_group(matches: list[re.Match]) -> list[re.Match]:
    for start, match in enumerate(matches):
        if match.group(1) != "1":
            continue
        picked = {match.group(1): match}
        for candidate in matches[start + 1:]:
            picked.setdefault(candidate.group(1), candidate)
            if len(picked) == 4:
                return sorted(picked.values(), key=lambda item: item.start())
    return []


QUESTION_STARTERS = {
    "a",
    "according",
    "after",
    "an",
    "as",
    "based",
    "before",
    "during",
    "for",
    "following",
    "from",
    "how",
    "in",
    "one",
    "the",
    "these",
    "this",
    "to",
    "under",
    "what",
    "which",
    "why",
}


BAD_STEM_START = re.compile(
    r"^(?:ala|ark|calif|colo|conn|del|fla|ga|ill|ind|iowa|kans|ky|la|mass|md|me|mich|minn|miss|mo|neb|nev|n\.h|n\.j|n\.c|n\.y|ohio|oreg|pa|penna|r\.i|s\.c|tenn|tex|va|vt|w\.|wis|messidor|source|electoral|popular)\b",
    re.I,
)


def score_question_candidate(stem: str, choices: list[dict], marker_line: str) -> int:
    normalized = clean_display(stem)
    first_word = (re.findall(r"[A-Za-z][A-Za-z.'-]*", normalized[:80]) or [""])[0].lower().rstrip(".")
    score = 0
    if first_word in QUESTION_STARTERS:
        score += 28
    if re.search(r"\b(which|what|how|why|according|based|identify|infer|conclusion|claim|result|cause|effect)\b", normalized[:140], re.I):
        score += 14
    if "\t" in marker_line:
        score += 4
    if 18 <= len(normalized) <= 420:
        score += 12
    elif len(normalized) > 900:
        score -= 25
    if BAD_STEM_START.search(normalized):
        score -= 35
    if re.search(r"\bSource:|\bElectoral Votes\b|\bPopular\b", normalized, re.I):
        score -= 12
    if all(choice.get("text") and not re.fullmatch(r"Choice\s+[1-4]", choice["text"]) for choice in choices):
        score += 18
    if any(len(choice.get("text", "")) < 3 for choice in choices):
        score -= 20
    return score


def parse_question_candidate(page_text: str, number: int, match: re.Match) -> tuple[int, str, list[dict]] | None:
    line_end = page_text.find("\n", match.start())
    marker_line = page_text[match.start(): line_end if line_end != -1 else len(page_text)]
    tail = page_text[match.end():]
    matches = list(re.finditer(r"\(\s*([1-4])\s*\)", tail))
    picked = choice_group(matches)
    if len(picked) != 4:
        return None

    stem = clean_display(tail[:picked[0].start()])
    choices = []
    for index, choice_match in enumerate(picked):
        label = choice_match.group(1)
        start = choice_match.end()
        end = picked[index + 1].start() if index + 1 < len(picked) else len(tail)
        if index == len(picked) - 1:
            next_marker = any_question_marker().search(tail, pos=start)
            if next_marker:
                end = min(end, next_marker.start())
        choice_text = clean_display(trim_choice_tail(tail[start:end]))
        choices.append({"label": label, "text": choice_text})
    choices = sorted(choices, key=lambda choice: int(choice["label"]))

    if not stem or len(choices) != 4 or any(not choice["text"] for choice in choices):
        return None
    return score_question_candidate(stem, choices, marker_line), stem, choices


def extract_question_from_pages(texts: list[str], number: int, part_two_page: int | None) -> tuple[int, str, list[dict]]:
    search_end = max(2, (part_two_page or min(len(texts), 18)) - 1)
    candidates: list[tuple[int, int, str, list[dict]]] = []
    for page_number in range(2, min(search_end, len(texts)) + 1):
        page_text = texts[page_number - 1]
        for match in question_marker(number).finditer(page_text):
            parsed = parse_question_candidate(page_text, number, match)
            if parsed:
                score, stem, choices = parsed
                candidates.append((score, page_number, stem, choices))
    if not candidates:
        raise RuntimeError(f"could not digitize MCQ {number}: no stem with four choices found")
    score, page_number, stem, choices = max(candidates, key=lambda item: item[0])
    if score < 20:
        raise RuntimeError(f"could not confidently digitize MCQ {number}: best candidate on page {page_number} scored {score}")
    return page_number, stem, choices


def question_has_stimulus(page_text: str, number: int) -> bool:
    pattern = re.compile(r"Base your answers?\s+to\s+questions?\s+(\d{1,2})(?:\s*(?:and|through|to|-)\s*(\d{1,2}))?", re.I)
    for match in pattern.finditer(page_text.replace("–", "-")):
        start = int(match.group(1))
        end = int(match.group(2) or start)
        if start <= number <= end:
            return True
    return False


def prompt_for_label(texts: list[str], pages: list[int], label: str) -> str:
    joined = "\n".join(texts[page - 1] for page in pages if 1 <= page <= len(texts))
    match = re.search(rf"(?m)^\s*{re.escape(label)}\s+(.+?)\s+\[\d\]", joined, flags=re.S)
    if not match:
        return ""
    return clean_display(match.group(1))


def seq_task_prompts(texts: list[str], start: int, end: int, question_number: int) -> list[str]:
    joined = "\n".join(texts[page - 1] for page in range(start, min(end, len(texts) + 1)))
    task_match = re.search(r"Task:\s*(.+?)(?=\n\s*(?:In developing|Document 1|SEQ Set)|$)", joined, flags=re.I | re.S)
    if not task_match:
        return [f"Write Short Essay Question {question_number} using the two released documents."]
    body = task_match.group(1)
    bullets = [clean_display(item) for item in re.findall(r"•\s*(.+?)(?=(?:\n\s*•|\n\s*In developing|\n\s*Document 1|$))", body, flags=re.S)]
    if bullets:
        return bullets
    prompt = clean_display(body)
    return [prompt] if prompt else [f"Write Short Essay Question {question_number} using the two released documents."]


def normalize_us_scaffold_titles(tasks: list[dict]) -> list[dict]:
    normalized = []
    for index, task in enumerate(tasks, start=1):
        item = dict(task)
        if re.search(r"\bscaffold\b", item.get("title", ""), re.I):
            item["title"] = f"Civic SAQ {30 + index}"
        normalized.append(item)
    return normalized


def civic_fallback_prompt(doc_number: int) -> str:
    prompts = {
        1: "Based on Document 1, identify one historical circumstance related to the civic issue.",
        2: "Based on Document 2, describe one effort by individuals, groups, or government to address the civic issue.",
        3: "Based on Document 3, explain one argument or point of view about the civic issue.",
        4: "Based on Document 4, describe one impact of an effort to address the civic issue.",
        5: "Based on Document 5, identify one limitation, success, or consequence of an effort to address the civic issue.",
        6: "Based on Document 6, explain how the civic issue continued or changed over time.",
    }
    return prompts.get(doc_number, f"Based on Document {doc_number}, answer the Civic Literacy short-answer question with one specific, supported response.")


def build_mcq(course: str, administration: str, form_id: str, pdf: fitz.Document, texts: list[str], out_dir: Path, keys: dict[int, str], part_two_page: int | None) -> list[dict]:
    items = []
    for number in range(1, 29):
        page_number, stem, choices = extract_question_from_pages(texts, number, part_two_page)
        asset = page_asset(pdf, out_dir, page_number, f"{administration} official exam page {page_number}")
        items.append(
            {
                "id": f"{form_id}-mcq-{number:02d}",
                "course": course,
                "set": f"{administration} Official Released Exam",
                "day": administration,
                "number": number,
                "officialQuestionNumber": number,
                "source": f"{administration} official exam page {page_number}",
                "stem": stem,
                "choices": choices,
                "correct": keys[number],
                "explanation": f"The official NYSED scoring key lists choice {keys[number]} for question {number}. Review the attached official page image if you missed it.",
                "stimulusRequired": question_has_stimulus(texts[page_number - 1], number),
                "sourceIntegrity": "trusted-official-page-image",
                "stimulusImages": [asset],
                "tags": ["official released exam", administration, COURSE_META[course]["label"], *keywords(stem + " " + " ".join(choice["text"] for choice in choices), 4)],
            }
        )
    return items


def generic_global_writing(course: str, administration: str, pdf: fitz.Document, texts: list[str], out_dir: Path) -> tuple[list[dict], list[dict], dict]:
    crq1 = find_page(texts, [r"CRQ\s+Set\s+1"], start=2) or find_page(texts, [r"Part\s+II", r"SHORT-ANSWER"], start=2) or 18
    crq2 = find_page(texts, [r"CRQ\s+Set\s+2"], crq1 + 1) or crq1 + 3
    essay_start = find_page(texts, [r"ENDURING\s+ISSUES\s+ESSAY"], crq2 + 1) or find_page(texts, [r"Part\s+III"], crq2 + 1) or crq2 + 3
    planning = find_page(texts, [r"OPTIONAL\s+PLANNING\s+PAGE"], essay_start + 1) or len(texts) + 1
    crq1_pages = doc_pages(texts, crq1, crq2, 2)
    crq2_pages = doc_pages(texts, crq2, essay_start, 2)
    essay_pages = doc_pages(texts, essay_start, planning, 5)

    def docs_for(section: str, pages: list[int], start_number: int) -> list[dict]:
        docs = []
        for offset, page_number in enumerate(pages):
            number = start_number + offset
            asset = page_asset(pdf, out_dir, page_number, f"{administration} {section} official page {page_number}")
            docs.append(make_doc(course, administration, section, number, page_number, asset, texts[page_number - 1]))
        return docs

    short_tasks = [
        {
            "id": "short-1",
            "title": "CRQ Set 1",
            "points": 3,
            "docs": docs_for("CRQ Set 1", crq1_pages, 1),
            "prompts": [prompt_for_label(texts, range(crq1, crq2), label) or fallback for label, fallback in [
                ("29", "Explain the historical circumstances surrounding Document 1."),
                ("30", "Explain the sourcing skill for Document 2."),
                ("31", "Identify and explain a relationship between Documents 1 and 2."),
            ]],
            "answerKey": ["Use the official rating guide for this administration. A supported response must answer each prompt and use the attached document evidence."],
            "modelAnswer": "A top practice response states the correct historical context, explains the sourcing skill when asked, and names a document-supported relationship using evidence from both CRQ Set 1 documents.",
        },
        {
            "id": "short-2",
            "title": "CRQ Set 2",
            "points": 4,
            "docs": docs_for("CRQ Set 2", crq2_pages, 3),
            "prompts": [prompt_for_label(texts, range(crq2, essay_start), label) or fallback for label, fallback in [
                ("32", "Explain the historical or geographic context for Document 1."),
                ("33", "Explain the sourcing skill for Document 2."),
                ("34a", "Identify a relationship or turning point using Documents 1 and 2."),
                ("34b", "Explain that relationship or turning point using evidence from both documents."),
            ]],
            "answerKey": ["Use the official rating guide for this administration. Credit comes from specific context, relationship or turning-point analysis, and evidence from the attached documents."],
            "modelAnswer": "A top practice response explains the setting for each document, identifies the required relationship or turning point, and uses evidence from both CRQ Set 2 documents to explain significance.",
        },
    ]
    essay_docs = docs_for("Enduring Issues Essay", essay_pages, 1)
    essay = {
        "id": "essay",
        "title": "Part III: Document-Based Enduring Issues Essay",
        "points": 5,
        "docMinimum": 3,
        "docs": essay_docs,
        "prompt": COURSE_META[course]["essayPrompt"],
        "answerKey": "Strong essays identify one enduring issue, define it, use at least three of the five official documents, explain significance, show endurance or change over time, and add outside information.",
        "modelEssay": "A top practice essay makes a clear enduring-issue claim, develops at least three document examples with explanation, connects the issue across time or place, adds accurate outside information, and stays organized around the same issue.",
    }
    return short_tasks, [], essay


def generic_us_writing(course: str, administration: str, pdf: fitz.Document, texts: list[str], out_dir: Path) -> tuple[list[dict], list[dict], dict]:
    seq_start = find_page(texts, [r"SHORT-ESSAY\s+QUESTIONS|SHORT ESSAY QUESTIONS|SEQs"], start=2) or 17
    seq2 = find_page(texts, [r"Set\s+2"], seq_start + 1) or seq_start + 3
    civic = find_page(texts, [r"CIVIC\s+LITERACY\s+ESSAY"], seq2 + 1) or seq2 + 3
    part_b = find_page(texts, [r"Part\s+B", r"Civic\s+Literacy\s+Essay\s+Question"], civic + 1) or len(texts) + 1
    seq1_pages = doc_pages(texts, seq_start, seq2, 2)
    seq2_pages = doc_pages(texts, seq2, civic, 2)
    civic_pages = doc_pages(texts, civic, part_b, 6)

    def docs_for(section: str, pages: list[int], start_number: int) -> list[dict]:
        docs = []
        for offset, page_number in enumerate(pages):
            number = start_number + offset
            asset = page_asset(pdf, out_dir, page_number, f"{administration} {section} official page {page_number}")
            docs.append(make_doc(course, administration, section, number, page_number, asset, texts[page_number - 1]))
        return docs

    seq1_docs = docs_for("Short Essay Set 1", seq1_pages, 1)
    seq2_docs = docs_for("Short Essay Set 2", seq2_pages, 3)
    civic_docs = docs_for("Civic Literacy", civic_pages, 1)
    short_tasks = [
        {
            "id": "short-1",
            "title": "Short Essay 1",
            "points": 5,
            "docs": seq1_docs,
            "prompts": seq_task_prompts(texts, seq_start, seq2, 29),
            "answerKey": ["Use the official rating guide for this administration. A strong response explains historical context, document relationship, evidence, and outside information."],
            "modelAnswer": "A top practice short essay explains the historical context, states the required relationship between the two documents, supports it with evidence from both documents, and includes accurate outside information.",
        },
        {
            "id": "short-2",
            "title": "Short Essay 2",
            "points": 5,
            "docs": seq2_docs,
            "prompts": seq_task_prompts(texts, seq2, civic, 30),
            "answerKey": ["Use the official rating guide for this administration. A strong response answers the exact skill prompt shown on the page and uses both documents."],
            "modelAnswer": "A top practice short essay answers the exact Set 2 prompt, uses both official documents, explains rather than summarizes, and adds an accurate U.S. History detail beyond the documents.",
        },
    ]
    scaffold_tasks = []
    for index, doc in enumerate(civic_docs, start=1):
        scaffold_tasks.append(
            {
                "id": f"scaffold-{index}",
                "title": f"Civic SAQ {30 + index}",
                "points": 1,
                "docs": [doc],
                "prompt": prompt_for_label(texts, civic_pages, str(30 + index)) or civic_fallback_prompt(index + 1),
                "answerKey": ["One specific, accurate answer supported by this document earns the practice point."],
                "modelAnswer": "A credited practice response gives one specific fact from the document and connects it to the civic issue.",
            }
        )
    essay = {
        "id": "essay",
        "title": "Part IIIB: Civic Literacy Essay",
        "points": 5,
        "docMinimum": 4,
        "docs": civic_docs,
        "prompt": COURSE_META[course]["essayPrompt"],
        "answerKey": "Strong essays identify the civic or constitutional issue, explain historical circumstances, describe efforts to address the issue, discuss the extent of success, use at least four documents, and add outside information.",
        "modelEssay": "A top practice civic literacy essay states the civic issue, explains the historical circumstances, develops at least two efforts to address the issue, evaluates success or limits, uses at least four official documents, and includes accurate outside information.",
    }
    return short_tasks, scaffold_tasks, essay


def load_existing_rich_forms() -> dict[tuple[str, str], dict]:
    if not OUT.exists():
        return {}
    try:
        existing = json.loads(OUT.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    rich = {}
    for course, form in (existing.get("forms") or {}).items():
        administration = form.get("administration")
        if administration:
            rich[(course, administration)] = form
    return rich


def build_form(course: str, exam: dict, rich_forms: dict[tuple[str, str], dict]) -> dict:
    meta = COURSE_META[course]
    form_id = exam["id"]
    administration = exam["administration"]
    form_slug = f"{meta['assetPrefix']}-{slug(administration)}"
    out_dir = ASSET_ROOT / form_slug
    pdf_path = download(exam["examUrl"], TMP / form_slug / url_filename(exam["examUrl"], f"{form_slug}-exam.pdf"))
    key_path = download(exam["scoringKeyExcelUrl"], TMP / form_slug / url_filename(exam["scoringKeyExcelUrl"], f"{form_slug}-sk.xlsx"))
    chart_path = download(exam["conversionChartExcelUrl"], TMP / form_slug / url_filename(exam["conversionChartExcelUrl"], f"{form_slug}-cc.xlsx"))

    keys = extract_scoring_key(key_path)
    conversion = extract_conversion_chart(chart_path)
    pdf = fitz.open(pdf_path)
    texts = page_texts(pdf)
    part_two_page = find_page(texts, [r"Part\s+II", r"SHORT-ANSWER|SHORT-ESSAY|SHORT ESSAY|SEQs"], start=2)
    mcq = build_mcq(course, administration, form_id, pdf, texts, out_dir, keys, part_two_page)

    if meta["profileId"] == "global-history-ii":
        short_tasks, scaffold_tasks, essay = generic_global_writing(course, administration, pdf, texts, out_dir)
    else:
        short_tasks, scaffold_tasks, essay = generic_us_writing(course, administration, pdf, texts, out_dir)

    rich = rich_forms.get((course, administration))
    if rich and administration == "January 2026":
        short_tasks = rich.get("shortTasks") or short_tasks
        scaffold_tasks = rich.get("scaffoldTasks") or scaffold_tasks
        essay = rich.get("essay") or essay
    if meta["profileId"] == "us-history":
        scaffold_tasks = normalize_us_scaffold_titles(scaffold_tasks)

    pdf.close()
    return {
        "id": form_id,
        "course": course,
        "profileId": meta["profileId"],
        "administration": administration,
        "mode": "exact-released-form",
        "officialExamUrl": exam["examUrl"],
        "largeTypeExamUrl": exam.get("largeTypeExamUrl", ""),
        "scoringKeyUrl": exam.get("scoringKeyPdfUrl") or exam.get("scoringKeyExcelUrl"),
        "ratingGuideUrl": (exam.get("ratingGuideUrls") or [{}])[0].get("url", ""),
        "ratingGuideUrls": exam.get("ratingGuideUrls") or [],
        "conversionChartUrl": exam.get("conversionChartPdfUrl") or exam.get("conversionChartExcelUrl"),
        "conversionChartExcelUrl": exam.get("conversionChartExcelUrl", ""),
        "conversionTable": conversion,
        "mcqNumbers": list(range(1, 29)),
        "mcq": mcq,
        "shortTasks": short_tasks,
        "scaffoldTasks": scaffold_tasks,
        "essay": essay,
    }


def main() -> int:
    catalog = json.loads(CATALOG_PATH.read_text(encoding="utf-8"))
    rich_forms = load_existing_rich_forms()
    forms_by_course: dict[str, list[dict]] = {}
    for course, data in catalog["courses"].items():
        forms = []
        for exam in data.get("exams", []):
            if not (exam.get("examUrl") and exam.get("scoringKeyExcelUrl") and exam.get("conversionChartExcelUrl")):
                continue
            print(f"Importing {course} {exam['administration']}")
            forms.append(build_form(course, exam, rich_forms))
        forms_by_course[course] = forms

    latest_forms = {course: forms[0] for course, forms in forms_by_course.items() if forms}
    payload = {
        "version": "20260430-interactive-past-regents-bank",
        "generatedAt": now_iso(),
        "sourceNote": "Practice-only interactive released forms built from official NYSED past examinations, scoring-key Excel files, rating guides, and conversion charts. MCQ pages are exact official page images; writing is graded by a local practice rubric and labeled as an estimate.",
        "forms": latest_forms,
        "formsByCourse": forms_by_course,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    for course, forms in forms_by_course.items():
        print(f"{course}: {len(forms)} interactive forms")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
